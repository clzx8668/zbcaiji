"""站点清单入库与 Excel 导出脚本（独立运行，无需项目依赖）。

用途：
  1. 将聚合招标站、三桶油/能源化工平台、东部十省市政府采购网写入 site_configs；
  2. 更新既有站点（中国政府采购网/山东政府采购网）的关键词；
  3. 导出全量站点清单 Excel。

运行：
  python scripts/seed_sites.py
"""

import sqlite3
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "bid_scraper.db"
EXPORT_PATH = ROOT / "data" / "output" / f"站点清单_{date.today().strftime('%Y%m%d')}.xlsx"

# (站点名称, 站点地址, 搜索类型, 关键词, 回望天数, 搜索URL模板, Cron, 是否启用, 备注)
# 关键词为逗号分隔；当前爬虫每次只用第一个关键词搜索，其余留给结果过滤。
SITES = [
    # ── 聚合型招标信息网站（免费关键词搜索）──
    ("全国公共资源交易平台", "https://www.ggzy.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "国家级公共资源交易总门户，免费；搜索为 JS/POST，需站点适配"),
    ("中国招标投标公共服务平台", "http://www.cebpubservice.com/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "国家级免费聚合平台，公告全；搜索交互复杂，需站点适配"),
    ("采招网", "https://www.bidcenter.com.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "商业聚合站，免费关键词搜索"),
    ("中国采购与招标网", "https://www.chinabidding.com.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "商业聚合站，免费关键词搜索"),
    ("比地招标网", "https://www.bidizhaobiao.com/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "商业聚合站，首页有 JS 防护，需站点适配"),

    # ── 三桶油 + 能源化工集团 ──
    ("中国石油招标投标网", "https://www.cnpcbidding.com/", "both", "陶瓷膜,水处理,盐湖提锂,磷酸铁锂", 7, "", "0 9 * * *", 1,
     "中石油官方招标平台；门户页为 JS 壳，需适配"),
    ("中国石化电子招投标平台", "https://ec.sinopec.com/", "both", "陶瓷膜,水处理,盐湖提锂,磷酸铁锂", 7, "", "0 9 * * *", 1,
     "中石化物资/工程招投标平台；bidding.sinopec.com 探测超时，改用 ec.sinopec.com"),
    ("中海油采办业务管理与交易系统", "https://buy.cnooc.com.cn/", "both", "陶瓷膜,水处理,盐湖提锂,磷酸铁锂", 7, "", "0 9 * * *", 1,
     "中海油官方采办平台；探测曾返回 50x，需确认可用性"),
    ("山东能源集团招标投标交易平台", "https://snzb.minegoods.com/", "both", "陶瓷膜,水处理,盐湖提锂,磷酸铁锂", 7, "", "0 9 * * *", 1,
     "山东能源集团官方平台"),
    ("中国华能集团电子商务平台", "https://ec.chng.com.cn/", "both", "陶瓷膜,水处理,盐湖提锂,磷酸铁锂", 7, "", "0 9 * * *", 1,
     "华能官方电子商务平台；有反爬（412），需适配"),
    ("中煤招标与采购网", "http://www.zmzb.com/", "both", "陶瓷膜,水处理,盐湖提锂,磷酸铁锂", 7, "", "0 9 * * *", 1,
     "中煤集团招标平台；门户为 JS 跳转，需适配"),
    ("国家能源招标网", "https://www.chnenergybidding.com.cn/bidweb", "both", "陶瓷膜,水处理,盐湖提锂,磷酸铁锂", 7, "", "0 9 * * *", 1,
     "国家能源集团官方招标网；首页自动跳转 /bidweb"),

    # ── 东部十省市政府采购网 ──
    ("北京市政府采购网", "http://www.ccgp-beijing.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1, ""),
    ("上海政府采购网", "http://www.ccgp-shanghai.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1, ""),
    ("天津市政府采购网", "http://www.ccgp-tianjin.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1, ""),
    ("江苏政府采购网", "http://www.ccgp-jiangsu.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1, ""),
    ("浙江政府采购网", "https://zfcg.czt.zj.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "SPA 站点，需适配"),
    ("山东政府采购网", "https://www.ccgp-shandong.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * 1-5", 1,
     "Vue SPA，走 :8087 API + 验证码，适配中"),
    ("广东省政府采购网", "https://gdgpo.czt.gd.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "SPA 站点，需适配"),
    ("福建省政府采购网", "https://zfcg.czt.fujian.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "SPA 站点，需适配"),
    ("河北省政府采购网", "http://www.ccgp-hebei.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "首页按地区 cookie 跳转，需适配"),
    ("辽宁省采购网", "http://www.ccgp-liaoning.gov.cn/", "both", "陶瓷膜,水处理", 7, "", "0 9 * * *", 1,
     "SPA 站点，需适配"),
]

# 既有站点：只更新关键词/备注/启用状态，保留原 search_url、days_back 等
UPDATE_EXISTING = {
    "中国政府采购网": "陶瓷膜,水处理",
}


def upsert_sites(conn: sqlite3.Connection) -> dict:
    stats = {"inserted": 0, "updated": 0}
    for site_name, site_url, search_type, keywords, days_back, search_url, cron, enabled, notes in SITES:
        row = conn.execute("SELECT id FROM site_configs WHERE site_name=?", (site_name,)).fetchone()
        if row:
            conn.execute(
                """UPDATE site_configs SET site_url=?, search_type=?, keywords=?, enabled=?,
                   notes=?, updated_at=datetime('now') WHERE site_name=?""",
                (site_url, search_type, keywords, int(enabled), notes, site_name),
            )
            stats["updated"] += 1
        else:
            conn.execute(
                """INSERT INTO site_configs
                   (site_name, site_url, search_type, keywords, days_back, search_url,
                    cron_expr, enabled, notes)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (site_name, site_url, search_type, keywords, days_back, search_url,
                 cron, int(enabled), notes),
            )
            stats["inserted"] += 1

    for site_name, keywords in UPDATE_EXISTING.items():
        cur = conn.execute(
            "UPDATE site_configs SET keywords=?, updated_at=datetime('now') WHERE site_name=?",
            (keywords, site_name),
        )
        if cur.rowcount:
            stats["updated"] += cur.rowcount
    return stats


def export_excel(conn: sqlite3.Connection, path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = conn.execute(
        """SELECT site_name, site_url, search_type, keywords, days_back, search_url,
                  cron_expr, enabled, notes
           FROM site_configs ORDER BY enabled DESC, site_name"""
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "站点清单"
    headers = ["序号", "站点名称", "站点地址", "搜索类型", "关键词", "回望天数",
               "搜索URL模板", "定时表达式", "是否启用", "备注"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, r in enumerate(rows, 1):
        ws.append([
            idx,
            r["site_name"],
            r["site_url"],
            r["search_type"],
            r["keywords"],
            r["days_back"],
            r["search_url"],
            r["cron_expr"],
            "是" if r["enabled"] else "否",
            r["notes"],
        ])

    widths = [6, 26, 48, 10, 32, 10, 60, 14, 10, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return len(rows)


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        stats = upsert_sites(conn)
        conn.commit()
        total = conn.execute("SELECT COUNT(*) FROM site_configs").fetchone()[0]
        enabled = conn.execute("SELECT COUNT(*) FROM site_configs WHERE enabled=1").fetchone()[0]
        count = export_excel(conn, EXPORT_PATH)
        print(f"站点入库完成：新增 {stats['inserted']}，更新 {stats['updated']}，库内共 {total} 个（启用 {enabled}）")
        print(f"Excel 已导出：{EXPORT_PATH}（{count} 行）")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
